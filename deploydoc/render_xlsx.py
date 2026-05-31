"""Renders the Change Request Excel by walking every cell of the company template
and rendering any Jinja placeholder it finds (typically in column B) against the
manifest. Cell styles are preserved because only `.value` is rewritten.
"""
from __future__ import annotations

from jinja2 import Environment, StrictUndefined, UndefinedError
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Jinja delimiters identical to docxtpl, so placeholders look the same everywhere.
_ENV = Environment(undefined=StrictUndefined, autoescape=False)
# `| lines` joins a list with real newlines (Jinja string escapes are unreliable).
_ENV.filters["lines"] = lambda items: "\n".join(str(i) for i in (items or []))


def _flatten_context(manifest: dict) -> dict:
    """Expose change_request fields at top level so placeholders stay short:
    {{ business_justification }} rather than {{ change_request.business_justification }}.
    """
    cr = dict(manifest.get("change_request", {}))
    repos = cr.get("repos", [])
    ctx = dict(cr)
    ctx["change_request"] = cr  # still allow the qualified form
    ctx.setdefault("ticket_ids", _join_ids([t["id"] for t in cr.get("tickets", [])]))
    ctx.setdefault("repo_names", ", ".join(r["name"] for r in repos))
    # Pre-joined multiline helpers (one repo per line) for backout/deploy version cells.
    ctx.setdefault("backout_versions", "\n".join(f"{r['name']}: {r['prod_tag']}" for r in repos))
    ctx.setdefault("deploy_versions", "\n".join(f"{r['name']}: {r['deploy_tag']}" for r in repos))
    return ctx


def _join_ids(ids: list[str]) -> str:
    if not ids:
        return ""
    if len(ids) == 1:
        return ids[0]
    return ", ".join(ids[:-1]) + " && " + ids[-1]


def render_change_request(manifest: dict, template_path: str, out_path: str) -> list[str]:
    """Returns a list of human-readable warnings (e.g. placeholders that failed)."""
    ctx = _flatten_context(manifest)
    wb = load_workbook(template_path)
    warnings: list[str] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or "{{" not in val and "{%" not in val:
                    continue
                try:
                    rendered = _ENV.from_string(val).render(**ctx)
                except UndefinedError as e:
                    warnings.append(f"{ws.title}!{cell.coordinate}: {e.message}")
                    rendered = val  # leave the placeholder so the gap is visible
                cell.value = rendered
                if isinstance(rendered, str) and "\n" in rendered:
                    cur = cell.alignment or Alignment()
                    cell.alignment = Alignment(
                        wrap_text=True,
                        horizontal=cur.horizontal,
                        vertical=cur.vertical or "top",
                    )

    wb.save(out_path)
    return warnings
